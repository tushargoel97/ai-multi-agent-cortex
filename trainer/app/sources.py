"""Training-data sources — upload/register, extract text, chunk.

Admins can supply base training data as PDF files, Excel workbooks, website
links (any HTML page, .aspx included), or free-text prompts. Files and the
manifest live under ``<data_dir>/sources/``.
"""

from __future__ import annotations

import json
import logging
import re
import time
import uuid
from pathlib import Path

from .config import settings

logger = logging.getLogger(__name__)

ALLOWED_UPLOAD_EXTS = {".pdf", ".xlsx", ".xls", ".png", ".jpg", ".jpeg", ".webp"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
_IMAGE_MIME = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".webp": "image/webp",
}

_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)


def _sources_dir() -> Path:
    d = settings.data_dir / "sources"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _manifest_path() -> Path:
    return _sources_dir() / "sources.json"


def list_sources() -> list[dict]:
    path = _manifest_path()
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text())
    except Exception:  # noqa: BLE001
        logger.exception("Corrupt sources manifest — starting empty")
        return []


def _save_manifest(entries: list[dict]) -> None:
    _manifest_path().write_text(json.dumps(entries, indent=2))


def add_file_source(filename: str, content: bytes) -> dict:
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise ValueError(
            f"Unsupported file type {ext!r} — allowed: {sorted(ALLOWED_UPLOAD_EXTS)}"
        )
    sid = uuid.uuid4().hex[:12]
    stored = _sources_dir() / f"{sid}{ext}"
    stored.write_bytes(content)
    if ext in IMAGE_EXTS:
        kind = "image"
    elif ext in {".xlsx", ".xls"}:
        kind = "excel"
    else:
        kind = "pdf"
    entry = {
        "id": sid,
        "type": kind,
        "name": filename,
        "path": stored.name,
        "size_kb": round(len(content) / 1024, 1),
        "added_at": time.time(),
    }
    _save_manifest([*list_sources(), entry])
    return entry


def add_url_source(url: str) -> dict:
    if not re.match(r"^https?://", url):
        raise ValueError("URL must start with http:// or https://")
    entry = {
        "id": uuid.uuid4().hex[:12],
        "type": "url",
        "name": url,
        "url": url,
        "added_at": time.time(),
    }
    _save_manifest([*list_sources(), entry])
    return entry


def add_prompt_source(text: str, name: str | None = None) -> dict:
    if not text.strip():
        raise ValueError("Prompt text is empty")
    sid = uuid.uuid4().hex[:12]
    stored = _sources_dir() / f"{sid}.txt"
    stored.write_text(text)
    entry = {
        "id": sid,
        "type": "prompt",
        "name": name or (text.strip()[:48] + ("…" if len(text.strip()) > 48 else "")),
        "path": stored.name,
        "size_kb": round(len(text) / 1024, 1),
        "added_at": time.time(),
    }
    _save_manifest([*list_sources(), entry])
    return entry


def delete_source(source_id: str) -> bool:
    entries = list_sources()
    keep = [e for e in entries if e["id"] != source_id]
    removed = next((e for e in entries if e["id"] == source_id), None)
    if removed is None:
        return False
    if removed.get("path"):
        (_sources_dir() / removed["path"]).unlink(missing_ok=True)
    _save_manifest(keep)
    return True


# ── Extraction ───────────────────────────────────────────────────────────────


def extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages)


def _extract_xls(path: Path) -> str:
    import xlrd

    book = xlrd.open_workbook(str(path))
    parts: list[str] = []
    for sheet in book.sheets():
        parts.append(f"# Sheet: {sheet.name}")
        if sheet.nrows == 0:
            continue
        header = [str(sheet.cell_value(0, c)) or f"col{c}" for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows):
            cells = [
                f"{header[c]}: {sheet.cell_value(r, c)}"
                for c in range(sheet.ncols)
                if str(sheet.cell_value(r, c)).strip()
            ]
            if cells:
                parts.append("; ".join(cells))
    return "\n".join(parts)


def extract_excel(path: Path) -> str:
    if path.suffix.lower() == ".xls":
        return _extract_xls(path)
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        parts.append(f"# Sheet: {ws.title}")
        if header is None:
            continue
        header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
        for row in rows:
            cells = [
                f"{header[i]}: {v}"
                for i, v in enumerate(row)
                if v is not None and i < len(header)
            ]
            if cells:
                parts.append("; ".join(cells))
    wb.close()
    return "\n".join(parts)


def extract_url(url: str) -> str:
    """Readable text of a web page — Firecrawl (JS/anti-bot) first, else a
    plain HTML fetch. Large cap so the page can be chunked for Q&A."""
    from . import search

    return search.fetch_page(url, max_chars=40000)


def research_prompt(topic: str, on_log=None) -> str:
    """Treat a free-text prompt as a research topic: search the web and fetch
    the top results as source content for Q&A generation.

    Falls back to the literal prompt text when no search provider is configured
    or nothing usable is found, so it is never worse than before.
    """
    from . import search

    log = on_log or (lambda _m: None)
    topic = topic.strip()
    if not topic:
        return ""
    if not search.search_configured():
        log(
            "  no web-search key set (FIRECRAWL_API_KEY / BRAVE_API_KEY / …) — "
            "using the prompt text literally"
        )
        return topic
    query = topic if len(topic) <= 200 else topic[:200]
    log(f"  researching topic: {query}")
    hits = search.provider_search(f"{query} specifications", max_results=6)
    if not hits:
        log("  no search results — using the prompt text literally")
        return topic
    parts: list[str] = []
    for hit in hits:
        url = hit.get("url", "")
        if not url:
            continue
        try:
            text = search.fetch_page(url, max_chars=6000)
            if text and len(text) > 200:
                parts.append(f"[{hit.get('title', '')}]\n{text}")
                log(f"  read: {url[:80]}")
        except Exception as e:  # noqa: BLE001
            log(f"  skip {url[:60]}: {type(e).__name__}")
    if not parts:
        log("  fetched no usable content — using the prompt text literally")
        return topic
    return f"Topic: {topic}\n\n" + "\n\n".join(parts)


def extract_image(path: Path) -> str:
    """Transcribe a document image (spec sheet, screenshot, table) to text via
    the vision-capable QA model. Point TRAINER_QA_* at a model that accepts
    images (e.g. OpenAI gpt-4o) — the default local 1B model can't read images.
    """
    from .qa_generator import transcribe_image

    mime = _IMAGE_MIME.get(path.suffix.lower(), "image/png")
    return transcribe_image(path.read_bytes(), mime)


def extract_source(entry: dict, on_log=None) -> str:
    """Extract plain text from a manifest entry."""
    kind = entry["type"]
    if kind == "url":
        return extract_url(entry["url"])
    path = _sources_dir() / entry["path"]
    if kind == "pdf":
        return extract_pdf(path)
    if kind == "excel":
        return extract_excel(path)
    if kind == "image":
        return extract_image(path)
    if kind == "prompt":
        return research_prompt(path.read_text(), on_log=on_log)
    raise ValueError(f"Unknown source type {kind!r}")


